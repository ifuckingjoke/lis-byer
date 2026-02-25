from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from trgb import green

class Tables:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self._create_titles()

    def _create_titles(self):
        titles = ["Название", "Минимальная цена (₽)", "FN\n(offers)", "MW\n(offers)",
                  "FT\n(offers)", "WW\n(offers)", "BS\n(offers)", "Cтатус", "Ссылка"
                ]
        if self.ws:
            self.ws.append(titles)
    
    def add_row(self, name, min_price, factory_new, minimal_wear, field_tested, well_worn, battle_scarred, skin_status, link):
        if self.ws:
            self.ws.append([name, min_price, factory_new, minimal_wear, field_tested, well_worn, battle_scarred, skin_status, link])

            for column in range(1, self.ws.max_column + 1):
                col = get_column_letter(column)
                self.ws.column_dimensions[col].width = 25
            
            for cell in self.ws[1]:
                cell.alignment = Alignment(wrap_text=True)
        if self.ws:
            self.ws.row_dimensions[1].height = 35
    
    def save(self, filename):
        self.wb.save(filename)
        print(green("[ script ] Таблица успешно создана"))